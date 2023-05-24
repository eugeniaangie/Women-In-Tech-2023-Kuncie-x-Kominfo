from openpyxl import Workbook
from openpyxl.writer.excel import save_workbook

data_dummy = [{
        "nama_item" : "sepatu",
        "harga" : 2000000,
        "brand" : "ventela"
    },
    {
        "nama_item" : "tas",
        "harga" : 2000000,
        "brand" : "ck"
    },
    {
        "nama_item" : "celana",
        "harga" : 500,
        "brand" : "cenem"
    }]


wb  = Workbook()
ws_export = wb.active
ws_export.title = "report"

header_export = ['Name', 'Price', 'Brand']

for idx, val in enumerate(header_export):
    ws_export.cell(row=1, column=idx+1, value=val)
    
r=1
for item in data_dummy:
    r+=1
    ws_export.cell(row=r, column=1, value=item['nama_item'])
    ws_export.cell(row=r, column=2, value=item['harga'])
    ws_export.cell(row=r, column=3, value=item['brand'])
    
file_path = "C:/Users/eugenia.salsabillah/Documents/just my stuff/WIT/GettingStarted/new_file3.xlsx"
wb.save(file_path)